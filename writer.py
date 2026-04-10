"""Write extracted values into the Peer Data template, preserving formulas.

Key principles:
- Load with openpyxl (NOT data_only) so formulas are preserved as strings
- Write only into Input cells per the dictionary; never touch helper/formula cells
- Append rows below the header block (rows 15+) for each ticker
- Also write to the Src Docs sheet (ticker, company id, name, FYE, DEF14A date+link)
"""
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
from openpyxl.utils import column_index_from_string
import openpyxl

from schema import Field, load_fields

PEER_DATA_SHEET = 'Peer Data'
SRC_DOCS_SHEET = 'Src Docs'
FIRST_DATA_ROW = 15  # row 14 = "Peer Data" label, 15+ = ticker rows


@dataclass
class CompanyExtraction:
    ticker: str            # e.g. "NYSE:ACN"
    company_id: str        # CIK or other id
    company_name: str
    fiscal_year_end: Any   # datetime
    filing_date: Any       # datetime — DEF 14A date
    filing_url: str
    values: Dict[str, Any] # variable_key -> extracted value (None = leave blank)


def write_company_row(ws_peer, ws_src, row_idx: int, extraction: CompanyExtraction,
                      fields: List[Field]) -> None:
    """Write one company's data into Peer Data row + Src Docs row."""
    # --- Peer Data sheet ---
    # Col B = Ticker (with row-suffix per template convention, but we use plain ticker here)
    ws_peer.cell(row=row_idx, column=2, value=extraction.ticker)
    # Col C = Company name
    ws_peer.cell(row=row_idx, column=3, value=extraction.company_name)
    # Col D = FYE
    ws_peer.cell(row=row_idx, column=4, value=extraction.fiscal_year_end)
    # All other input columns from the dictionary
    for f in fields:
        val = extraction.values.get(f.key)
        if val is None or val == '':
            continue
        col_idx = column_index_from_string(f.col)
        # Don't overwrite if it's already a formula in the template's input row
        existing = ws_peer.cell(row=row_idx, column=col_idx).value
        if isinstance(existing, str) and existing.startswith('='):
            continue
        ws_peer.cell(row=row_idx, column=col_idx, value=val)

    # --- Src Docs sheet ---
    # Header in row 2: Ticker, Company ID, Company Name, FYE, DEF 14A Date, DEF 14A Link
    src_row = _next_empty_src_row(ws_src)
    ws_src.cell(row=src_row, column=1, value=extraction.ticker)
    ws_src.cell(row=src_row, column=2, value=extraction.company_id)
    ws_src.cell(row=src_row, column=3, value=extraction.company_name)
    ws_src.cell(row=src_row, column=4, value=extraction.fiscal_year_end)
    ws_src.cell(row=src_row, column=5, value=extraction.filing_date)
    ws_src.cell(row=src_row, column=6, value=extraction.filing_url)


def _next_empty_src_row(ws_src) -> int:
    r = 3
    while ws_src.cell(row=r, column=1).value not in (None, ''):
        r += 1
    return r


def populate_template(template_path: str, output_path: str,
                      extractions: List[CompanyExtraction], fields: List[Field]) -> None:
    wb = openpyxl.load_workbook(template_path)  # keep formulas
    ws_peer = wb[PEER_DATA_SHEET]
    ws_src = wb[SRC_DOCS_SHEET]
    for i, ex in enumerate(extractions):
        write_company_row(ws_peer, ws_src, FIRST_DATA_ROW + i, ex, fields)
    wb.save(output_path)
