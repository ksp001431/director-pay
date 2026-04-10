"""Round-trip test: prove the column mapping + writer are correct.

Reads ACN/ADP/CRM from the populated example file, writes them into a fresh
copy of the blank template via the writer, then cell-by-cell diffs the
result against the original example file.

If the diff is empty (for the columns we control), the writer is verified.
"""
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from schema import load_fields
from writer import populate_template, CompanyExtraction, FIRST_DATA_ROW

DICT = '/mnt/user-data/uploads/Peer_Data_Field_Dictionary_v1.xlsx'
EXAMPLE = '/mnt/user-data/uploads/DIRECTOR_PAY__Template-Example_.xlsx'
BLANK = '/mnt/user-data/uploads/DIRECTOR_PAY__Template_.xlsx'
OUT = '/home/claude/director_pay/roundtrip_output.xlsx'


def extract_row(ws, row_idx, fields):
    """Read values keyed by variable_key from a populated Peer Data row."""
    out = {}
    for f in fields:
        c = column_index_from_string(f.col)
        v = ws.cell(row=row_idx, column=c).value
        if v is not None and v != '':
            out[f.key] = v
    return out


def main():
    fields = load_fields(DICT)

    # Step 1: read the example file (data_only=True so we get values not formulas)
    ex_wb = openpyxl.load_workbook(EXAMPLE, data_only=True)
    ex_ws = ex_wb['Peer Data']
    ex_src = ex_wb['Src Docs']

    # Build extractions for the 3 example companies (rows 15, 16, 17)
    extractions = []
    for row_idx in (15, 16, 17):
        ticker = ex_ws.cell(row=row_idx, column=2).value
        company = ex_ws.cell(row=row_idx, column=3).value
        fye = ex_ws.cell(row=row_idx, column=4).value
        # Find filing info on Src Docs by ticker
        filing_date, filing_url, company_id = None, '', ''
        for sr in range(3, 50):
            if ex_src.cell(row=sr, column=1).value == ticker:
                company_id = ex_src.cell(row=sr, column=2).value or ''
                filing_date = ex_src.cell(row=sr, column=5).value
                filing_url = ex_src.cell(row=sr, column=6).value or ''
                break
        values = extract_row(ex_ws, row_idx, fields)
        extractions.append(CompanyExtraction(
            ticker=ticker, company_id=str(company_id), company_name=company,
            fiscal_year_end=fye, filing_date=filing_date, filing_url=filing_url,
            values=values,
        ))
        print(f"Extracted {ticker}: {len(values)} non-blank input fields")

    # Step 2: populate a fresh copy of the BLANK template
    populate_template(BLANK, OUT, extractions, fields)
    print(f"Wrote {OUT}")

    # Step 3: diff cell-by-cell against the example
    out_wb = openpyxl.load_workbook(OUT, data_only=False)  # preserve formulas
    out_ws = out_wb['Peer Data']
    # Reload example without data_only so formulas show as formulas, matching out
    ex_wb2 = openpyxl.load_workbook(EXAMPLE, data_only=False)
    ex_ws2 = ex_wb2['Peer Data']

    # Diff only the columns we control (input columns from the dictionary)
    # Plus B, C, D (ticker, company, fye)
    controlled_cols = {column_index_from_string(f.col) for f in fields}
    controlled_cols.update({2, 3, 4})

    diffs = []
    for row_idx in (15, 16, 17):
        for c in sorted(controlled_cols):
            v_out = out_ws.cell(row=row_idx, column=c).value
            v_ex = ex_ws2.cell(row=row_idx, column=c).value
            # Treat None == '' as equal
            if (v_out in (None, '')) and (v_ex in (None, '')):
                continue
            if v_out != v_ex:
                diffs.append((row_idx, get_column_letter(c), v_ex, v_out))

    print(f"\n=== ROUND-TRIP DIFF: {len(diffs)} differences ===")
    for d in diffs[:30]:
        print(f"  row {d[0]} col {d[1]}: example={d[2]!r}  out={d[3]!r}")
    if len(diffs) > 30:
        print(f"  ...and {len(diffs)-30} more")

    return len(diffs)


if __name__ == '__main__':
    import sys
    sys.exit(0 if main() == 0 else 1)
