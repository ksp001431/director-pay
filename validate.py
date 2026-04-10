"""Compare an LLM-generated output file against the known-correct example
to measure extraction accuracy field by field.

Usage:
  python validate.py --generated test_output.xlsx
"""
import argparse
import openpyxl
from openpyxl.utils import column_index_from_string
from collections import Counter
from schema import load_fields

EXAMPLE = "/mnt/user-data/uploads/DIRECTOR_PAY__Template-Example_.xlsx"
DICT = "/mnt/user-data/uploads/Peer_Data_Field_Dictionary_v1.xlsx"


def normalize(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str):
        return v.strip().lower()
    return v


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--generated", required=True)
    args = p.parse_args()

    fields = load_fields(DICT)
    ex = openpyxl.load_workbook(EXAMPLE, data_only=True)["Peer Data"]
    gen = openpyxl.load_workbook(args.generated, data_only=True)["Peer Data"]

    # Build ticker -> row index for both
    def row_map(ws):
        out = {}
        for r in range(15, 60):
            t = ws.cell(row=r, column=2).value
            if t:
                out[t] = r
        return out
    ex_rows = row_map(ex)
    gen_rows = row_map(gen)

    common = set(ex_rows) & set(gen_rows)
    print(f"Comparing {len(common)} ticker(s): {sorted(common)}\n")

    overall = Counter()
    for ticker in sorted(common):
        er, gr = ex_rows[ticker], gen_rows[ticker]
        match = mismatch = ex_only = gen_only = both_blank = 0
        details = []
        for f in fields:
            c = column_index_from_string(f.col)
            ev = normalize(ex.cell(row=er, column=c).value)
            gv = normalize(gen.cell(row=gr, column=c).value)
            if ev is None and gv is None:
                both_blank += 1
            elif ev is None and gv is not None:
                gen_only += 1
                details.append(("GEN_ONLY ", f.col, f.key, ev, gv))
            elif ev is not None and gv is None:
                ex_only += 1
                details.append(("MISSED   ", f.col, f.key, ev, gv))
            elif ev == gv:
                match += 1
            else:
                mismatch += 1
                details.append(("MISMATCH ", f.col, f.key, ev, gv))

        non_blank_in_ex = match + mismatch + ex_only
        accuracy = match / non_blank_in_ex * 100 if non_blank_in_ex else 0
        print(f"=== {ticker} ===")
        print(f"  Match: {match} | Mismatch: {mismatch} | Missed: {ex_only} | "
              f"Hallucinated: {gen_only} | Both blank: {both_blank}")
        print(f"  Recall on disclosed fields: {accuracy:.1f}%")
        for d in details[:15]:
            print(f"    {d[0]} {d[1]:4} {d[2][:40]:40} ex={d[3]!r}  gen={d[4]!r}")
        if len(details) > 15:
            print(f"    ...{len(details)-15} more")
        print()
        overall["match"] += match
        overall["mismatch"] += mismatch
        overall["missed"] += ex_only
        overall["hallucinated"] += gen_only

    total_disclosed = overall["match"] + overall["mismatch"] + overall["missed"]
    print(f"OVERALL across {len(common)} tickers:")
    print(f"  Disclosed fields in example: {total_disclosed}")
    print(f"  Correct: {overall['match']} ({overall['match']/total_disclosed*100:.1f}%)")
    print(f"  Wrong value: {overall['mismatch']}")
    print(f"  Missed: {overall['missed']}")
    print(f"  Hallucinated: {overall['hallucinated']}")


if __name__ == "__main__":
    main()
