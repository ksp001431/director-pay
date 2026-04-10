"""End-to-end orchestrator: ticker list -> populated xlsx with audit trail.

Adds:
- Hidden _Extraction Log sheet (one row per field per company)
- Yellow fill on low-confidence cells in Peer Data
- _Run Summary sheet listing skipped tickers and errors
"""
from typing import List, Dict, Any
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string

from schema import Field, load_fields
from writer import (CompanyExtraction, populate_template, FIRST_DATA_ROW,
                    PEER_DATA_SHEET, SRC_DOCS_SHEET)
from edgar import (find_director_pay_filing, fetch_filing_text,
                   SkippedForeignFiler, TickerNotFound, NoEligibleFiling)
from extractor import extract_from_filing

LOW_CONF_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
HEADER_FONT = Font(bold=True)


@dataclass
class RunResult:
    ticker: str
    status: str        # "ok", "skipped_foreign", "not_found", "no_filing", "error"
    detail: str = ""
    extraction: CompanyExtraction = None
    audit: List[Dict[str, Any]] = None


def run_one(ticker: str, fields: List[Field], provider: str, api_key: str,
            model: str = None) -> RunResult:
    try:
        filing = find_director_pay_filing(ticker)
    except SkippedForeignFiler as e:
        return RunResult(ticker, "skipped_foreign", str(e))
    except TickerNotFound as e:
        return RunResult(ticker, "not_found", str(e))
    except NoEligibleFiling as e:
        return RunResult(ticker, "no_filing", str(e))
    except Exception as e:
        return RunResult(ticker, "error", f"EDGAR lookup failed: {e}")

    try:
        html = fetch_filing_text(filing)
        values, audit = extract_from_filing(
            fields, html, ticker, filing.company_name, provider, api_key, model)
    except Exception as e:
        return RunResult(ticker, "error", f"Extraction failed: {e}")

    extraction = CompanyExtraction(
        ticker=ticker, company_id=filing.cik, company_name=filing.company_name,
        fiscal_year_end=values.get("fiscal_year_end_of_pay") or filing.fiscal_year_end,
        filing_date=filing.filing_date, filing_url=filing.url, values=values,
    )
    return RunResult(ticker, "ok", filing.form, extraction, audit)


def run_batch(tickers: List[str], template_path: str, output_path: str,
              provider: str, api_key: str, model: str = None,
              progress_cb=None) -> List[RunResult]:
    import os
    dict_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "Peer_Data_Field_Dictionary_v1.xlsx")
    fields = load_fields(dict_path)
    results: List[RunResult] = []
    for i, t in enumerate(tickers):
        if progress_cb:
            progress_cb(i, len(tickers), t)
        results.append(run_one(t, fields, provider, api_key, model))
    if progress_cb:
        progress_cb(len(tickers), len(tickers), "writing output")

    successful = [r for r in results if r.status == "ok"]
    extractions = [r.extraction for r in successful]
    populate_template(template_path, output_path, extractions, fields)

    # Add audit log + run summary + low-confidence highlighting
    wb = openpyxl.load_workbook(output_path)
    _write_audit_log(wb, successful, fields)
    _write_run_summary(wb, results)
    _highlight_low_confidence(wb, successful, fields)
    wb.save(output_path)
    return results


def _write_audit_log(wb, results: List[RunResult], fields: List[Field]):
    if "_Extraction Log" in wb.sheetnames:
        del wb["_Extraction Log"]
    ws = wb.create_sheet("_Extraction Log")
    ws.sheet_state = "hidden"
    headers = ["Ticker", "Variable Key", "Excel Col", "Block",
               "Value", "Confidence", "Section", "Source Quote"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = HEADER_FONT
    fmap = {f.key: f for f in fields}
    row = 2
    for r in results:
        for entry in (r.audit or []):
            f = fmap.get(entry["key"])
            ws.cell(row=row, column=1, value=r.ticker)
            ws.cell(row=row, column=2, value=entry["key"])
            ws.cell(row=row, column=3, value=f.col if f else "")
            ws.cell(row=row, column=4, value=f.block if f else "")
            ws.cell(row=row, column=5, value=entry.get("value"))
            ws.cell(row=row, column=6, value=entry.get("confidence"))
            ws.cell(row=row, column=7, value=entry.get("section"))
            ws.cell(row=row, column=8, value=entry.get("quote"))
            row += 1
    for col, w in [("A",12),("B",40),("C",8),("D",30),("E",18),("F",12),("G",30),("H",60)]:
        ws.column_dimensions[col].width = w


def _write_run_summary(wb, results: List[RunResult]):
    if "_Run Summary" in wb.sheetnames:
        del wb["_Run Summary"]
    ws = wb.create_sheet("_Run Summary", 0)  # first tab
    headers = ["Ticker", "Status", "Detail / Form"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = HEADER_FONT
    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r.ticker)
        ws.cell(row=i, column=2, value=r.status)
        ws.cell(row=i, column=3, value=r.detail)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60


def _highlight_low_confidence(wb, results: List[RunResult], fields: List[Field]):
    ws = wb[PEER_DATA_SHEET]
    fmap = {f.key: f for f in fields}
    for i, r in enumerate(results):
        row_idx = FIRST_DATA_ROW + i
        for entry in (r.audit or []):
            if entry.get("confidence") == "low" and entry.get("value") not in (None, ""):
                f = fmap.get(entry["key"])
                if not f:
                    continue
                col_idx = column_index_from_string(f.col)
                ws.cell(row=row_idx, column=col_idx).fill = LOW_CONF_FILL
